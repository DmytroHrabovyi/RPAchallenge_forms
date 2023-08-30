import openpyxl
import os


class DataManager:
    @staticmethod
    def get_file_path():
        directory_path = 'C:/Users/User/Downloads'
        most_recent_file = None
        most_recent_time = 0

        for entry in os.scandir(directory_path):
            if entry.is_file():
                mod_time = entry.stat().st_mtime_ns
                if mod_time > most_recent_time:
                    most_recent_file = entry.name
                    most_recent_time = mod_time

        return directory_path + '/' + most_recent_file

    @staticmethod
    def extract_data(path):
        dataframe = openpyxl.load_workbook(path)
        dataframe1 = dataframe.active

        data_dict = {}
        for col in dataframe1.iter_cols():
            if col[0].value is not None:
                data_dict[col[0].value.strip()] = []

        for col in dataframe1.iter_cols():
            if col is None:
                break
            for row in range(1, dataframe1.max_row):
                if col[row].value is None:
                    break
                data_dict[col[0].value.strip()].append(col[row].value)

        return data_dict
