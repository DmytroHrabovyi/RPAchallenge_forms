from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import time
import os


class Driver:
    def __init__(self, driver, website):
        self.driver = driver
        self.driver.get(website)
        self.driver.maximize_window()

    def download_data(self):
        download_link = self.driver.find_element(By.XPATH, '/html/body/app-root/div[2]/app-rpa1/div/div[1]/div[6]/a')
        download_link.click()
        time.sleep(5)

    def get_file(self):
        self.download_data()

        directory_path = 'C:/Users/User/Downloads'
        most_recent_file = None
        most_recent_time = 0

        for entry in os.scandir(directory_path):
            if entry.is_file():
                mod_time = entry.stat().st_mtime_ns
                if mod_time > most_recent_time:
                    most_recent_file = entry.name
                    most_recent_time = mod_time

        return directory_path + '/' + most_recent_file

    def extract_data(self):
        dataframe = openpyxl.load_workbook(self.get_file())
        dataframe1 = dataframe.active
        data_list = []
        for row in range(1, dataframe1.max_row):
            for col in dataframe1.iter_cols():
                if col[row].value is not None:
                    data_list.append(col[row].value)

        return data_list

    def upload_data(self):
        data = self.extract_data()

        start_button = self.driver.find_element(By.XPATH, '/html/body/app-root/div[2]/app-rpa1/div/div[1]/div[6]/button')
        start_button.click()

        paths = [
            'labelFirstName',
            'labelLastName',
            'labelCompanyName',
            'labelRole',
            'labelAddress',
            'labelEmail',
            'labelPhone'
        ]

        while data:
            for path in paths:
                field = self.driver.find_element(By.XPATH, f'//*[@ng-reflect-name="{path}"]')
                field.click()
                field.send_keys(data.pop(0))

            submit_button = self.driver.find_element(By.CLASS_NAME, 'btn.uiColorButton')
            submit_button.click()

        time.sleep(5)


driver = Driver(webdriver.Chrome(), 'https://rpachallenge.com/?lang=EN')
driver.upload_data()

